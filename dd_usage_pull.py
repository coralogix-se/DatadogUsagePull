#!/usr/bin/env python3
"""
dd_usage_pull.py — Datadog → Coralogix Usage Report
=====================================================
Pulls Datadog usage metrics from the official Usage Metering API, mirrors
every tile on the Bill Overview page, converts the numbers to Coralogix
sizing using the same formulas as the sizing Excel template, then packages
all outputs into a single ZIP file ready to hand off.

Usage
-----
    python dd_usage_pull.py [--month YYYY-MM] [--site datadoghq.com] [--out DIR]

Environment variables (put these in a .env file next to the script):
    DD_API_KEY   Datadog API key          (needs usage_read)
    DD_APP_KEY   Datadog Application key  (needs usage_read)
    DD_SITE      datadoghq.com | datadoghq.eu | us3/us5.datadoghq.com  (default: datadoghq.com)
    DD_MONTH     YYYY-MM  (default: previous calendar month)

This script collects usage volumes only (logs, metrics, traces, RUM, synthetics).
It does NOT call Datadog cost or billing-dollar endpoints.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sys
import time
import zipfile
from dataclasses import dataclass, field, fields as dc_fields
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

# ── Dependency guards ──────────────────────────────────────────────────────
try:
    import requests
except ImportError:
    sys.exit("\n  Missing 'requests'. Run:  pip install requests\n")

try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit("\n  Missing 'python-dotenv'. Run:  pip install python-dotenv\n")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("  [warn] openpyxl not installed — Excel output will be skipped. Run: pip install openpyxl")

load_dotenv()

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — Constants  (all sizing assumptions live here)
# ═══════════════════════════════════════════════════════════════════════════

AVG_LOG_SIZE_KB  = 2.5     # assumed average compressed log line size
AVG_SPAN_SIZE_KB = 1.5     # assumed average span payload size
TS_PER_HOST      = 750     # estimated Prometheus/DD time series per host or container
TS_TO_UNITS      = 3.3e-5  # Coralogix metrics: TimeSeries → Units/day conversion
DAYS_PER_MONTH   = 30.0    # used for all monthly → daily conversions
HOURS_PER_MONTH  = 30.0 * 24  # 720 — Datadog _sum fields for hosts/containers are in host-hours
SW_LABEL_FACTOR  = 0.30    # serverless TS = invocations × 3 labels × 10% unique = ×0.30
LOG_TIER_MON     = 0.70    # Monitoring share of ingested logs
LOG_TIER_COMP    = 0.30    # Compliance share
SPAN_TIER_MON    = 0.10    # Monitoring share of ingested spans
SPAN_TIER_COMP   = 0.90    # Compliance share of ingested spans

KNOWN_SITES: dict[str, str] = {
    "datadoghq.com":     "https://api.datadoghq.com",
    "datadoghq.eu":      "https://api.datadoghq.eu",
    "us3.datadoghq.com": "https://api.us3.datadoghq.com",
    "us5.datadoghq.com": "https://api.us5.datadoghq.com",
    "ddog-gov.com":      "https://api.ddog-gov.com",
}

FRESHNESS_NOTE = (
    "Datadog usage data may be delayed up to 72 hours."
)

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — Datadog API client
# ═══════════════════════════════════════════════════════════════════════════

class DatadogClient:
    """Thin, retry-aware wrapper around Datadog's Usage Metering API."""

    MAX_RETRIES   = 4
    RETRY_WAIT_S  = 2  # exponential: 2, 4, 8 seconds

    def __init__(self, api_key: str, app_key: str, site: str = "datadoghq.com"):
        if site not in KNOWN_SITES:
            sys.exit(
                f"\n  Unknown Datadog site: '{site}'\n"
                f"  Allowed values: {', '.join(KNOWN_SITES)}\n"
            )
        self.base_url = KNOWN_SITES[site]
        self.headers  = {
            "Accept":             "application/json",
            "DD-API-KEY":         api_key,
            "DD-APPLICATION-KEY": app_key,
        }

    def _get(self, path: str, params: dict | None = None) -> dict:
        url = f"{self.base_url}{path}"
        for attempt in range(self.MAX_RETRIES):
            try:
                resp = requests.get(url, headers=self.headers, params=params or {}, timeout=45)
                if resp.status_code == 429:
                    wait = self.RETRY_WAIT_S ** (attempt + 1)
                    print(f"    Rate limited — waiting {wait}s (retry {attempt+1}/{self.MAX_RETRIES})")
                    time.sleep(wait)
                    continue
                if resp.status_code == 403:
                    raise PermissionError(
                        f"403 Forbidden: {path}\n"
                        "  Ensure the API and Application keys have usage_read."
                    )
                if resp.status_code == 400:
                    raise ValueError(f"400 Bad Request: {path} — {resp.text[:300]}")
                resp.raise_for_status()
                return resp.json()
            except (requests.ConnectionError, requests.Timeout) as exc:
                if attempt == self.MAX_RETRIES - 1:
                    raise
                wait = self.RETRY_WAIT_S ** (attempt + 1)
                print(f"    Network error ({exc}) — retrying in {wait}s")
                time.sleep(wait)
        raise RuntimeError(f"All retries exhausted for {path}")

    # ── Individual endpoint wrappers ────────────────────────────────────────

    def usage_summary(self, start_month: str, end_month: str | None = None) -> dict:
        p: dict[str, str] = {"start_month": start_month}
        if end_month:
            p["end_month"] = end_month
        return self._get("/api/v1/usage/summary", p)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — Data model
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class UsageSnapshot:
    """All Datadog usage metrics for one calendar month."""

    month:        str   = ""
    account_name: str   = ""
    region:       str   = ""
    site:         str   = ""
    pulled_at:    str   = ""

    # ── Infrastructure ──────────────────────────────────────────────────────
    infra_hosts:    float = 0   # Infra Hosts tile (concurrent count)
    apm_hosts:      float = 0   # APM Hosts tile (concurrent count)
    containers:     float = 0   # concurrent container count (for sizing formulas)
    container_hours: float = 0  # Container Hours tile (total host-hours in month)
    network_hosts:  float = 0   # Network Hosts tile
    dbm_hosts:      float = 0   # DBM Hosts tile

    # ── Profiling ───────────────────────────────────────────────────────────
    profiled_hosts:      float = 0
    profiled_containers: float = 0
    profiled_fargate:    float = 0
    apm_fargate:         float = 0
    fargate_tasks:       float = 0

    # ── Custom Metrics ──────────────────────────────────────────────────────
    custom_metrics:          float = 0  # Custom Metrics tile (time series)
    ingested_custom_metrics: float = 0  # Ingested Custom Metrics tile

    # ── Logs ────────────────────────────────────────────────────────────────
    ingested_logs_bytes:      float = 0   # Ingested Logs tile (bytes)
    indexed_logs_3day:        float = 0   # events
    indexed_logs_7day:        float = 0
    indexed_logs_15day:       float = 0
    indexed_logs_30day:       float = 0
    indexed_logs_45day:       float = 0
    indexed_logs_60day:       float = 0
    indexed_logs_90day:       float = 0
    indexed_logs_180day:      float = 0
    indexed_logs_360day:      float = 0
    indexed_logs_live:        float = 0   # live search (short-term)
    indexed_logs_rehydrated:  float = 0   # rehydrated from archive
    security_logs_bytes:      float = 0   # SIEM analyzed logs

    # ── APM / Tracing ───────────────────────────────────────────────────────
    ingested_spans_bytes:  float = 0   # Ingested Spans tile (bytes)
    indexed_spans:         float = 0   # Indexed Spans tile (events)
    custom_events:         float = 0   # Custom Events tile

    # ── Serverless ──────────────────────────────────────────────────────────
    serverless_functions:     float = 0   # Serverless Workload Functions tile
    serverless_invocations:   float = 0   # total invocations (monthly)
    serverless_app_instances: float = 0   # Serverless App Instances tile

    # ── RUM ─────────────────────────────────────────────────────────────────
    rum_sessions:          float = 0   # RUM Investigate tile
    rum_lite_sessions:     float = 0   # RUM Measure tile
    rum_replay:            float = 0   # Session Replay tile
    rum_errors:            float = 0   # Error Tracking Events

    # ── Synthetics ──────────────────────────────────────────────────────────
    synthetics_api:     float = 0
    synthetics_browser: float = 0

    # ── Other ───────────────────────────────────────────────────────────────
    incident_management_seats:     float = 0
    test_optimization_committers:  float = 0
    test_optimization_spans:       float = 0
    product_analytics_sessions:    float = 0
    session_replay:                float = 0   # may alias rum_replay
    app_builder_apps:              float = 0
    bits_ai_investigations:        float = 0


    # ── Full API responses for raw JSON export ───────────────────────────────
    raw: dict = field(default_factory=dict)


@dataclass
class CoralogixSizing:
    """Coralogix sizing estimates derived from Datadog usage."""

    # ── Logs ────────────────────────────────────────────────────────────────
    total_ingested_logs_gb_month:  float = 0
    total_indexed_logs_count:      float = 0   # events
    indexed_logs_size_gb_month:    float = 0
    indexed_pct_logs:              float = 0
    daily_logs_gb:                 float = 0
    daily_logs_fs_gb:              float = 0   # unused, kept for compat
    daily_logs_mon_gb:             float = 0   # Monitoring       (40 %)
    daily_logs_comp_gb:            float = 0   # Compliance       (10 %)

    # ── Metrics ─────────────────────────────────────────────────────────────
    host_count:               float = 0
    container_count:          float = 0
    host_container_ts:        float = 0
    sw_func_ts:               float = 0
    sw_invoc_ts:              float = 0
    total_ts:                 float = 0
    metrics_units_per_day:    float = 0

    # ── Tracing ─────────────────────────────────────────────────────────────
    ingested_spans_gb_month:  float = 0
    indexed_spans_gb_month:   float = 0
    indexed_pct_spans:        float = 0
    daily_spans_ingest_gb:    float = 0
    daily_spans_mon_gb:       float = 0
    daily_spans_comp_gb:      float = 0
    daily_spans_indexed_gb:   float = 0
    daily_spans_archive_gb:   float = 0

    # ── RUM ─────────────────────────────────────────────────────────────────
    rum_sessions_monthly:        float = 0
    rum_sessions_daily:          float = 0
    rum_session_recording_daily: float = 0
    rum_errors_per_day:          float = 0

    # ── Synthetics → Checkly ────────────────────────────────────────────────
    synthetics_api_daily:        float = 0
    synthetics_browser_daily:    float = 0
    synthetics_api_monthly:      float = 0
    synthetics_browser_monthly:  float = 0


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — Extraction helpers
# ═══════════════════════════════════════════════════════════════════════════

def _f(item: dict, *keys: str) -> float:
    """Return the first non-None numeric value found among `keys`, else 0."""
    for k in keys:
        v = item.get(k)
        if v is not None:
            try:
                return float(v)
            except (TypeError, ValueError):
                continue
    return 0.0


def extract_usage_snapshot(
    summary_raw: dict,
    month: str,
    site: str,
) -> UsageSnapshot:
    """Build a UsageSnapshot from all available API responses."""

    snap = UsageSnapshot(
        month=month,
        site=site,
        pulled_at=datetime.now(timezone.utc).isoformat(),
    )

    # ── Usage summary ────────────────────────────────────────────────────────
    usage_list = summary_raw.get("usage", [])
    if usage_list:
        # Select the item whose date matches the target month, or use the last item.
        item = usage_list[-1]
        for u in usage_list:
            date_str = str(u.get("date", u.get("start_date", "")))
            if date_str.startswith(month):
                item = u
                break

        # Account name: try top-level first, then orgs[0]
        snap.account_name = item.get("account_name", item.get("org_name", ""))
        if not snap.account_name:
            orgs = item.get("orgs", [])
            if orgs:
                snap.account_name = orgs[0].get("account_name", orgs[0].get("org_name", ""))
        snap.region = item.get("region", "")

        # Infrastructure
        # IMPORTANT: Datadog's v2 API uses "_sum" fields that represent TOTAL HOST-HOURS
        # for the month (not a concurrent count). We divide by HOURS_PER_MONTH (720) to
        # recover the average concurrent host count for sizing purposes.
        # "top99p" fields are already a concurrent count (peak), so they're used as-is.

        def _hosts_from_hours(hours_field: str, *top99p_fields: str) -> float:
            """Try top99p (count) first; fall back to hours ÷ 720."""
            for k in top99p_fields:
                v = _f(item, k)
                if v: return v
            h = _f(item, hours_field)
            return h / HOURS_PER_MONTH if h else 0.0

        # Sum all host types; prefer top99p counts, fall back to hours/720
        snap.infra_hosts = (
            _hosts_from_hours("agent_host_sum",       "agent_host_top99p_sum",   "agent_host_top99p")
            + _hosts_from_hours("aws_host_sum",       "aws_host_top99p_sum",     "aws_host_top99p")
            + _hosts_from_hours("azure_host_sum",     "azure_host_top99p_sum",   "azure_host_top99p")
            + _hosts_from_hours("gcp_host_sum",       "gcp_host_top99p_sum",     "gcp_host_top99p")
            + _hosts_from_hours("vsphere_host_sum",   "vsphere_host_top99p_sum", "vsphere_host_top99p")
            + _hosts_from_hours("alibaba_host_sum",   "alibaba_host_top99p_sum")
            + _hosts_from_hours("heroku_host_sum",    "heroku_host_top99p_sum")
            + _hosts_from_hours("opentelemetry_host_sum", "opentelemetry_host_top99p_sum")
        ) or _hosts_from_hours("infra_hours_sum",     "infra_host_top99p_sum",   "infra_host_top99p")

        snap.apm_hosts = _hosts_from_hours(
            "apm_host_sum",
            "apm_host_top99p_sum", "apm_host_top99p",
        ) or _hosts_from_hours("apm_host_incl_usm_sum", "apm_host_incl_usm_top99p")

        snap.network_hosts = _f(item, "npm_host_top99p", "npm_host_top99p_sum") or \
            _hosts_from_hours("npm_host_sum", "network_device_count_top99p_sum")

        snap.dbm_hosts = _f(item, "dbm_host_top99p", "dbm_host_top99p_sum",
                            "dbm_host_database_instance_top99p")

        # Containers: "_sum" = total container-hours for the month; ÷720 = concurrent count
        raw_container_hours = _f(item, "container_sum", "container_count_avg_sum",
                                 "container_avg_sum")
        snap.container_hours = raw_container_hours   # used for the "Container Hours" tile
        snap.containers = (
            _f(item, "container_count_avg", "container_avg")   # already an average — use as-is
            or (raw_container_hours / HOURS_PER_MONTH if raw_container_hours else 0.0)
        )

        # Profiling / Fargate
        # profiling_host_top99p is already a concurrent count; profiling_container_agent_count_sum is hours
        snap.profiled_hosts = _f(item,
            "profiling_host_top99p",
            "profiling_host_count_top99p_sum", "profiling_host_count_top99p",
            "profiling_uncategorized_host_count_top99p",
        )
        _prof_cont_hours = _f(item, "profiling_container_agent_count_sum")
        snap.profiled_containers = (
            _f(item, "profiling_container_agent_count_avg", "profiling_container_count_avg_sum")
            or (_prof_cont_hours / HOURS_PER_MONTH if _prof_cont_hours else 0.0)
        )
        snap.profiled_fargate = _f(item,
            "avg_profiled_fargate_tasks",
            "avg_profiled_fargate_tasks_hw_max_sum",
            "profiling_aas_count_top99p_sum",
            "fargate_container_profiler_profiling_fargate_avg",
        )
        snap.apm_fargate   = _f(item, "apm_fargate_count_avg_sum", "apm_fargate_count_avg")
        snap.fargate_tasks = _f(item, "fargate_tasks_count_avg_sum",
                                "fargate_tasks_count_avg", "fargate_tasks_count_hwm")

        # Custom Metrics
        snap.custom_metrics          = _f(item, "custom_ts_avg", "custom_ts_avg_sum",
                                          "custom_timeseries_avg_sum")
        snap.ingested_custom_metrics = _f(item, "custom_ingested_timeseries_average_sum",
                                          "ingested_custom_timeseries_average_sum",
                                          "custom_live_ts_avg_sum", "custom_live_ts_avg")

        # Logs — ingested bytes
        # "live_ingested_bytes_sum" is the most common field in newer API responses
        snap.ingested_logs_bytes = _f(item,
            "live_ingested_bytes_sum",           # ← most common in 2025+ responses
            "ingested_events_bytes_sum",
            "ingested_events_bytes_agg_sum",
            "billable_ingested_bytes_agg_sum",
            "logs_live_ingested_bytes_agg_sum",
        )

        # Indexed logs — Datadog uses two naming conventions for retention tiers:
        #   logs_indexed_logs_usage_sum_N_day  (newer, e.g. 2025+)
        #   logs_indexed_Nday_agg_sum          (older)
        snap.indexed_logs_live       = _f(item,
            "logs_indexed_live_index_indexed_sum",
            "logs_live_indexed_logs_usage_sum",
            "logs_live_indexed_count_agg_sum",
            "live_indexed_events_sum",
        )
        snap.indexed_logs_rehydrated = _f(item,
            "logs_rehydrated_indexed_count_agg_sum",
            "rehydrated_indexed_events_sum",
        )
        snap.indexed_logs_3day   = _f(item,
            "logs_indexed_logs_usage_sum_3_day",   # newer naming
            "logs_indexed_3day_agg_sum",
            "logs_indexed_3_day_agg_sum",
        )
        snap.indexed_logs_7day   = _f(item,
            "logs_indexed_logs_usage_sum_7_day",
            "logs_indexed_7day_agg_sum",
            "logs_indexed_7_day_agg_sum",
        )
        snap.indexed_logs_15day  = _f(item,
            "logs_indexed_logs_usage_sum_15_day",
            "logs_indexed_15day_agg_sum",
            "logs_indexed_15_day_agg_sum",
            "logs_indexed_logs_indexed_15day_sum",
        )
        snap.indexed_logs_30day  = _f(item,
            "logs_indexed_logs_usage_sum_30_day",
            "logs_indexed_30day_agg_sum",
            "logs_indexed_30_day_agg_sum",
            "logs_indexed_logs_indexed_30day_sum",
        )
        snap.indexed_logs_45day  = _f(item,
            "logs_indexed_logs_usage_sum_45_day",
            "logs_indexed_45day_agg_sum",
            "logs_indexed_45_day_agg_sum",
            "logs_indexed_logs_indexed_45day_sum",
        )
        snap.indexed_logs_60day  = _f(item,
            "logs_indexed_logs_usage_sum_60_day",
            "logs_indexed_60day_agg_sum",
            "logs_indexed_60_day_agg_sum",
        )
        snap.indexed_logs_90day  = _f(item,
            "logs_indexed_logs_usage_sum_90_day",
            "logs_indexed_90day_agg_sum",
            "logs_indexed_90_day_agg_sum",
            "logs_indexed_logs_indexed_90day_sum",
        )
        snap.indexed_logs_180day = _f(item,
            "logs_indexed_logs_usage_sum_180_day",
            "logs_indexed_180day_agg_sum",
            "logs_indexed_180_day_agg_sum",
        )
        snap.indexed_logs_360day = _f(item,
            "logs_indexed_logs_usage_sum_360_day",
            "logs_indexed_360day_agg_sum",
            "logs_indexed_360_day_agg_sum",
        )
        # SIEM / security logs
        snap.security_logs_bytes = _f(item,
            "twol_ingested_events_bytes_sum",      # "2nd line" (SIEM) ingested bytes
            "twol_ingested_events_bytes_agg_sum",
            "siem_ingested_bytes_agg_sum",
            "siem_analyzed_logs_add_on_count_sum",
        )

        # APM / Tracing
        snap.ingested_spans_bytes = _f(item,
            "twol_ingested_events_bytes_sum",  # Tracing Without Limits — primary ingestion field
            "ingested_spans_bytes_agg_sum",
            "ingested_spans_bytes_sum",
        )
        # apm_ingest_gb_sum only covers overages above the 150 GB/APM-host/month free tier;
        # use it only as a last resort when no byte-level field found, and convert GB → bytes.
        if snap.ingested_spans_bytes == 0:
            _apm_overage_gb = _f(item, "apm_ingest_gb_sum", "apm_ingest_only_gb_sum")
            if _apm_overage_gb:
                snap.ingested_spans_bytes = _apm_overage_gb * 1e9

        snap.indexed_spans = _f(item,
            "trace_search_indexed_events_count_sum",   # newer naming
            "trace_search_indexed_events_count_agg_sum",
            "apm_span_custom_agg_sum",
            "indexed_events_count_sum",
        )
        snap.custom_events = _f(item,
            "custom_events_agg_sum", "custom_events_sum",
        )

        # Serverless
        snap.serverless_functions     = _f(item,
            "serverless_func_count_avg_sum", "serverless_func_avg_sum",
            "serverless_func_count_agg_sum",
        )
        snap.serverless_invocations   = _f(item,
            "lambda_invocations_count_agg_sum",
            "serverless_invocation_count_agg_sum",
            "aws_lambda_invocations_sum",
        )
        snap.serverless_app_instances = _f(item,
            "serverless_apps_total_count_hw_max_sum",
            "serverless_apps_azure_count_hw_max_sum",
        )

        # RUM
        # Total RUM Investigate = all session types summed
        rum_total = _f(item,
            "rum_total_session_count_sum",
            "rum_browser_and_mobile_session_count_sum",
            "rum_session_count_sum",
        )
        rum_lite   = _f(item,
            "rum_lite_session_count_sum",
            "rum_browser_lite_session_count_sum",
            "rum_lite_session_count_agg_sum",
            "rum_browser_lite_session_count_agg_sum",
        )
        rum_replay = _f(item,
            "rum_replay_session_count_sum",
            "rum_browser_replay_session_count_sum",
            "rum_replay_session_count_agg_sum",
            "session_replay_count_agg_sum",
        )
        rum_legacy = _f(item,
            "rum_browser_legacy_session_count_sum",
            "browser_legacy_session_count_sum",
        )
        # If the total session count field is missing, sum the parts
        snap.rum_sessions      = rum_total or (rum_lite + rum_replay + rum_legacy)
        snap.rum_lite_sessions = rum_lite
        snap.rum_replay        = rum_replay
        snap.session_replay    = rum_replay  # alias
        snap.rum_errors        = _f(item,
            "error_tracking_error_events_sum",   # newer naming
            "error_tracking_events_sum",
            "error_tracking_events_agg_sum",
            "total_error_tracking_events_agg_sum",
        )

        # Synthetics
        snap.synthetics_api     = _f(item,
            "synthetics_check_calls_count_agg_sum",
            "synthetics_check_calls_count_sum",
        )
        snap.synthetics_browser = _f(item,
            "synthetics_browser_check_calls_count_sum",   # newer naming
            "synthetics_browser_check_calls_count_agg_sum",
            "browser_check_calls_count_agg_sum",
        )

        # Incident / DBM / CI / Other
        snap.incident_management_seats    = _f(item,
            "incident_management_seats_hwm",
            "incident_management_monthly_active_users_hwm",
            "incident_management_monthly_active_users_hw_max_sum",
            "incident_management_monthly_active_users_hw_max",
        )
        snap.test_optimization_committers = _f(item,
            "ci_visibility_pipeline_committers_hwm",   # newer naming
            "ci_visibility_itsm_committers_hw_max_sum",
            "ci_visibility_committers_hw_max_sum",
        )
        snap.test_optimization_spans      = _f(item,
            "ci_pipeline_indexed_spans_sum",            # newer naming
            "ci_test_indexed_spans_agg_sum",
            "ci_test_indexed_spans_sum",
            "ci_visibility_test_indexed_spans_agg_sum",
        )
        snap.product_analytics_sessions   = _f(item,
            "product_analytics_sum",
            "product_analytics_count_agg_sum",
            "product_analytics_session_count_agg_sum",
        )
        snap.app_builder_apps    = _f(item, "published_app_hwm_sum", "published_app_hw_max_sum")
        snap.bits_ai_investigations = _f(item,
            "bits_ai_investigations_sum",
            "bits_ai_total_conversations_agg_sum",
            "bits_ai_investigations_agg_sum",
            "ai_credits_bits_sre_ai_credits_sum",
        )

    # ── Store raw for JSON export ────────────────────────────────────────────
    snap.raw = {
        "usage_summary": summary_raw,
    }

    return snap


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5 — Coralogix conversion  (matches the Excel template formulas)
# ═══════════════════════════════════════════════════════════════════════════

def compute_coralogix_sizing(snap: UsageSnapshot) -> CoralogixSizing:
    """Apply the sizing formulas from the Excel template to the usage snapshot."""

    cx = CoralogixSizing()

    # ── Logs ─────────────────────────────────────────────────────────────────
    ingested_gb = snap.ingested_logs_bytes / 1e9
    security_gb = snap.security_logs_bytes / 1e9
    cx.total_ingested_logs_gb_month = ingested_gb + security_gb

    # Indexed logs = live/short-term + rehydrated + 15-day + 90-day+ retention
    # Map: Live & Rehydrated = (3day + 7day + live + rehydrated)
    #      15-day tier = indexed_15day
    #      90-day+ tier = 30d + 45d + 60d + 90d + 180d + 360d
    indexed_live_rehydrated = (
        snap.indexed_logs_3day
        + snap.indexed_logs_7day
        + snap.indexed_logs_live
        + snap.indexed_logs_rehydrated
    )
    indexed_15d = snap.indexed_logs_15day
    indexed_long = (
        snap.indexed_logs_30day + snap.indexed_logs_45day
        + snap.indexed_logs_60day + snap.indexed_logs_90day
        + snap.indexed_logs_180day + snap.indexed_logs_360day
    )

    cx.total_indexed_logs_count = indexed_live_rehydrated + indexed_15d + indexed_long

    # Size in bytes: count × avg_log_size_kb × 1024 (bytes per KB)
    # Then convert bytes → GB: ÷ 1024³
    indexed_bytes = cx.total_indexed_logs_count * AVG_LOG_SIZE_KB * 1024
    cx.indexed_logs_size_gb_month = indexed_bytes / (1024 ** 3)

    if cx.total_ingested_logs_gb_month > 0:
        cx.indexed_pct_logs = cx.indexed_logs_size_gb_month / cx.total_ingested_logs_gb_month

    cx.daily_logs_gb      = cx.total_ingested_logs_gb_month / DAYS_PER_MONTH
    cx.daily_logs_mon_gb  = cx.daily_logs_gb * LOG_TIER_MON
    cx.daily_logs_comp_gb = cx.daily_logs_gb * LOG_TIER_COMP

    # ── Metrics ──────────────────────────────────────────────────────────────
    # Hosts = infra + apm + profiled + network + fargate types
    cx.host_count = (
        snap.infra_hosts + snap.apm_hosts + snap.profiled_hosts + snap.network_hosts
        + snap.fargate_tasks + snap.profiled_fargate + snap.apm_fargate
    )
    cx.container_count = snap.containers + snap.profiled_containers

    cx.host_container_ts = (cx.host_count + cx.container_count) * TS_PER_HOST

    # Serverless metrics: daily functions/invocations × 3 labels × 10% unique dimensions
    sw_func_daily  = snap.serverless_functions / DAYS_PER_MONTH
    sw_invoc_daily = snap.serverless_invocations / DAYS_PER_MONTH
    cx.sw_func_ts  = sw_func_daily  * SW_LABEL_FACTOR
    cx.sw_invoc_ts = sw_invoc_daily * SW_LABEL_FACTOR

    cx.total_ts           = cx.host_container_ts + snap.custom_metrics + cx.sw_func_ts + cx.sw_invoc_ts
    cx.metrics_units_per_day = cx.total_ts * TS_TO_UNITS

    # ── Tracing ──────────────────────────────────────────────────────────────
    cx.ingested_spans_gb_month = snap.ingested_spans_bytes / 1e9

    # Indexed spans GB = (count + custom_events) × avg_span_size_kb / 1024² (KB→GB)
    cx.indexed_spans_gb_month = (
        (snap.indexed_spans + snap.custom_events) * AVG_SPAN_SIZE_KB
    ) / (1024 * 1024)

    if cx.ingested_spans_gb_month > 0:
        cx.indexed_pct_spans = cx.indexed_spans_gb_month / cx.ingested_spans_gb_month

    cx.daily_spans_ingest_gb  = cx.ingested_spans_gb_month / DAYS_PER_MONTH
    cx.daily_spans_mon_gb     = cx.daily_spans_ingest_gb * SPAN_TIER_MON
    cx.daily_spans_comp_gb    = cx.daily_spans_ingest_gb * SPAN_TIER_COMP
    cx.daily_spans_indexed_gb = cx.daily_spans_ingest_gb * cx.indexed_pct_spans
    cx.daily_spans_archive_gb = cx.daily_spans_ingest_gb - cx.daily_spans_indexed_gb

    # ── RUM ──────────────────────────────────────────────────────────────────
    cx.rum_sessions_monthly        = snap.rum_sessions
    cx.rum_sessions_daily          = snap.rum_sessions / DAYS_PER_MONTH
    cx.rum_session_recording_daily = snap.session_replay / DAYS_PER_MONTH
    cx.rum_errors_per_day   = snap.rum_errors / DAYS_PER_MONTH

    # ── Synthetics → Checkly ─────────────────────────────────────────────────
    cx.synthetics_api_monthly     = snap.synthetics_api
    cx.synthetics_browser_monthly = snap.synthetics_browser
    cx.synthetics_api_daily       = snap.synthetics_api / DAYS_PER_MONTH
    cx.synthetics_browser_daily   = snap.synthetics_browser / DAYS_PER_MONTH

    return cx


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 6 — Formatting helpers
# ═══════════════════════════════════════════════════════════════════════════

def _fmt(n: float | None, decimals: int = 1, suffix: str = "") -> str:
    if n is None:
        return "N/A"
    if n == 0:
        return f"0{suffix}"
    if abs(n) >= 1e12:
        return f"{n/1e12:.{decimals}f}T{suffix}"
    if abs(n) >= 1e9:
        return f"{n/1e9:.{decimals}f}B{suffix}"
    if abs(n) >= 1e6:
        return f"{n/1e6:.{decimals}f}M{suffix}"
    if abs(n) >= 1e3:
        return f"{n/1e3:.{decimals}f}K{suffix}"
    return f"{n:.{decimals}f}{suffix}"

def _bytes_to_tb(b: float) -> str:
    return f"{b/1e12:.2f} TB"

def _bytes_to_gb(b: float) -> str:
    return f"{b/1e9:.2f} GB"


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 6b — Multi-month trend analysis
# ═══════════════════════════════════════════════════════════════════════════

def compute_trends(
    pairs: list[tuple["UsageSnapshot", "CoralogixSizing"]],
) -> list[dict]:
    """Compute month-over-month % changes for the key TCO metrics."""
    results: list[dict] = []
    for i, (snap, cx) in enumerate(pairs):
        entry: dict = {
            "month":           snap.month,
            "logs_gb_day":     cx.daily_logs_gb,
            "metrics_ts":      cx.total_ts,
            "tracing_gb_day":  cx.daily_spans_ingest_gb,
            "rum_day":         cx.rum_sessions_daily,
            "rum_rec_day":     cx.rum_session_recording_daily,
        }
        if i > 0:
            prev = results[i - 1]
            def _pct(curr: float, prev_val: float) -> float | None:
                if prev_val:
                    return (curr - prev_val) / abs(prev_val) * 100
                return None
            entry["logs_pct"]     = _pct(entry["logs_gb_day"],    prev["logs_gb_day"])
            entry["metrics_pct"]  = _pct(entry["metrics_ts"],     prev["metrics_ts"])
            entry["tracing_pct"]  = _pct(entry["tracing_gb_day"], prev["tracing_gb_day"])
            entry["rum_pct"]      = _pct(entry["rum_day"],         prev["rum_day"])
        else:
            entry["logs_pct"] = entry["metrics_pct"] = entry["tracing_pct"] = entry["rum_pct"] = None
        results.append(entry)
    return results


def _trend_arrow(pct: float | None) -> str:
    """Return a coloured HTML arrow for a % change."""
    if pct is None:
        return '<span style="color:#aaa">—</span>'
    if pct > 5:
        return f'<span style="color:#e04b2a">▲ {pct:+.1f}%</span>'
    if pct < -5:
        return f'<span style="color:#22a06b">▼ {pct:+.1f}%</span>'
    return f'<span style="color:#888">≈ {pct:+.1f}%</span>'


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 7 — CSV output
# ═══════════════════════════════════════════════════════════════════════════

def generate_csv(snap: UsageSnapshot, cx: CoralogixSizing) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Datadog Usage Report", snap.month])
    w.writerow(["Account", snap.account_name or "N/A"])
    w.writerow(["Site", snap.site])
    w.writerow(["Pulled at", snap.pulled_at])
    w.writerow(["Note", FRESHNESS_NOTE])
    w.writerow([])

    # ── Bill Overview tiles ───────────────────────────────────────────────
    w.writerow(["=== DATADOG BILL OVERVIEW ==="])
    w.writerow(["Metric", "Value", "Unit"])

    tiles = [
        ("Infra Hosts",                    snap.infra_hosts,          "hosts (avg concurrent)"),
        ("APM Hosts",                      snap.apm_hosts,            "hosts (avg concurrent)"),
        ("Custom Metrics",                 snap.custom_metrics,       "time series"),
        ("Ingested Custom Metrics",        snap.ingested_custom_metrics, "time series"),
        ("Indexed Logs (3 Day)",           snap.indexed_logs_3day,    "events"),
        ("Indexed Logs (7 Day)",           snap.indexed_logs_7day,    "events"),
        ("Indexed Logs (15 Day)",          snap.indexed_logs_15day,   "events"),
        ("Indexed Logs (30 Day)",          snap.indexed_logs_30day,   "events"),
        ("Indexed Logs (45 Day)",          snap.indexed_logs_45day,   "events"),
        ("Indexed Logs (60 Day)",          snap.indexed_logs_60day,   "events"),
        ("Indexed Logs (90 Day)",          snap.indexed_logs_90day,   "events"),
        ("Indexed Logs (180 Day)",         snap.indexed_logs_180day,  "events"),
        ("Indexed Logs (360 Day)",         snap.indexed_logs_360day,  "events"),
        ("Indexed Logs (Live Search)",     snap.indexed_logs_live,    "events"),
        ("Indexed Logs (Rehydrated)",      snap.indexed_logs_rehydrated, "events"),
        ("Ingested Logs",                  snap.ingested_logs_bytes,  "bytes"),
        ("Container Hours",                snap.container_hours,      "container-hours/month"),
        ("Containers (avg concurrent)",    snap.containers,           "containers"),
        ("Ingested Spans",                 snap.ingested_spans_bytes, "bytes"),
        ("Indexed Spans",                  snap.indexed_spans,        "events"),
        ("Profiled Hosts",                 snap.profiled_hosts,       "hosts"),
        ("Profiled Container Hours",       snap.profiled_containers,  "container-hours"),
        ("Serverless Workload Functions",  snap.serverless_functions, "functions"),
        ("Serverless Invocations",         snap.serverless_invocations, "invocations"),
        ("Serverless App Instances",       snap.serverless_app_instances, "instances"),
        ("Fargate Tasks",                  snap.fargate_tasks,        "tasks"),
        ("APM Fargate Tasks",              snap.apm_fargate,          "tasks"),
        ("Profiled Fargate Tasks",         snap.profiled_fargate,     "tasks"),
        ("Network Hosts",                  snap.network_hosts,        "hosts"),
        ("DBM Hosts",                      snap.dbm_hosts,            "hosts"),
        ("Synthetics API Test Runs",       snap.synthetics_api,       "test runs"),
        ("Synthetics Browser Test Runs",   snap.synthetics_browser,   "test runs"),
        ("RUM Investigate (Sessions)",     snap.rum_sessions,         "sessions"),
        ("RUM Measure (Lite Sessions)",    snap.rum_lite_sessions,    "sessions"),
        ("Session Replay",                 snap.rum_replay,           "sessions"),
        ("Error Tracking Events",          snap.rum_errors,           "events"),
        ("Incident Management Seats",      snap.incident_management_seats, "seats"),
        ("Custom Events",                  snap.custom_events,        "events"),
        ("Product Analytics Sessions",     snap.product_analytics_sessions, "sessions"),
        ("Test Optimization Committers",   snap.test_optimization_committers, "committers"),
        ("Test Optimization Spans",        snap.test_optimization_spans, "spans"),
        ("App Builder Published Apps",     snap.app_builder_apps,     "apps"),
        ("Bits AI SRE Investigations",     snap.bits_ai_investigations, "investigations"),
        ("SIEM/Security Logs",             snap.security_logs_bytes,  "bytes"),
    ]
    for name, value, unit in tiles:
        w.writerow([name, value, unit])

    w.writerow([])

    # ── Coralogix sizing ─────────────────────────────────────────────────
    w.writerow(["=== CORALOGIX SIZING ==="])
    w.writerow(["Assumption: avg log size (KB)",   AVG_LOG_SIZE_KB])
    w.writerow(["Assumption: avg span size (KB)",  AVG_SPAN_SIZE_KB])
    w.writerow(["Assumption: TS per host/container", TS_PER_HOST])
    w.writerow(["Assumption: TS-to-Units factor",  TS_TO_UNITS])
    w.writerow(["Assumption: days per month",       DAYS_PER_MONTH])
    w.writerow([])

    w.writerow(["-- Logs --"])
    w.writerow(["Total Ingested Logs (GB/month)",   f"{cx.total_ingested_logs_gb_month:.2f}"])
    w.writerow(["Total Indexed Logs (events/month)", f"{cx.total_indexed_logs_count:.0f}"])
    w.writerow(["Total Indexed Logs Size (GB/month)", f"{cx.indexed_logs_size_gb_month:.2f}"])
    w.writerow(["Indexed Percentage",               f"{cx.indexed_pct_logs*100:.2f}%"])
    w.writerow(["Daily Ingested Logs (GB/day)",     f"{cx.daily_logs_gb:.2f}"])
    w.writerow(["  Monitoring 70% (GB/day)",        f"{cx.daily_logs_mon_gb:.2f}"])
    w.writerow(["  Compliance 30% (GB/day)",        f"{cx.daily_logs_comp_gb:.2f}"])
    w.writerow([])

    w.writerow(["-- Metrics --"])
    w.writerow(["Host count (all types)",           f"{cx.host_count:.0f}"])
    w.writerow(["Container count",                  f"{cx.container_count:.0f}"])
    w.writerow(["Host+Container TimeSeries",        f"{cx.host_container_ts:.0f}"])
    w.writerow(["Serverless Functions TS",          f"{cx.sw_func_ts:.2f}"])
    w.writerow(["Serverless Invocations TS",        f"{cx.sw_invoc_ts:.2f}"])
    w.writerow(["Total TimeSeries (NumSeries)",     f"{cx.total_ts:.0f}"])
    w.writerow(["Metrics Units/day",               f"{cx.metrics_units_per_day:.2f}"])
    w.writerow([])

    w.writerow(["-- Tracing --"])
    w.writerow(["Ingested Spans (GB/month)",        f"{cx.ingested_spans_gb_month:.2f}"])
    w.writerow(["Indexed Spans (GB/month)",         f"{cx.indexed_spans_gb_month:.2f}"])
    w.writerow(["Indexed Span Percentage",          f"{cx.indexed_pct_spans*100:.4f}%"])
    w.writerow(["Daily Ingested Spans (GB/day)",    f"{cx.daily_spans_ingest_gb:.2f}"])
    w.writerow(["  Monitoring 10% (GB/day)",        f"{cx.daily_spans_mon_gb:.2f}"])
    w.writerow(["  Compliance 90% (GB/day)",        f"{cx.daily_spans_comp_gb:.2f}"])
    w.writerow(["  Indexed (GB/day)",               f"{cx.daily_spans_indexed_gb:.4f}"])
    w.writerow(["  Archive (GB/day)",               f"{cx.daily_spans_archive_gb:.2f}"])
    w.writerow([])

    w.writerow(["-- RUM --"])
    w.writerow(["RUM Sessions/month",               f"{cx.rum_sessions_monthly:.0f}"])
    w.writerow(["RUM Total Sessions/day",           f"{cx.rum_sessions_daily:.0f}"])
    w.writerow(["RUM Session Recording/day",        f"{cx.rum_session_recording_daily:.0f}"])
    w.writerow(["RUM Errors/day",                   f"{cx.rum_errors_per_day:.2f}"])
    w.writerow([])

    w.writerow(["-- Summary for TCO Calculator --"])
    w.writerow(["Logs GB/day",          f"{cx.daily_logs_gb:.2f}"])
    w.writerow(["Metrics NumSeries",    f"{cx.total_ts:.0f}"])
    w.writerow(["Tracing GB/day",       f"{cx.daily_spans_ingest_gb:.2f}"])
    w.writerow(["RUM Total Sessions/day",       f"{cx.rum_sessions_daily:.0f}"])
    w.writerow(["RUM Session Recording/day",    f"{cx.rum_session_recording_daily:.0f}"])
    w.writerow([])
    w.writerow(["-- Summary for Checkly --"])
    w.writerow(["API checks / day",       f"{cx.synthetics_api_daily:.0f}"])
    w.writerow(["Browser checks / day",   f"{cx.synthetics_browser_daily:.0f}"])
    w.writerow(["API checks / month",     f"{cx.synthetics_api_monthly:.0f}"])
    w.writerow(["Browser checks / month", f"{cx.synthetics_browser_monthly:.0f}"])

    return buf.getvalue().encode()


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 8 — Excel output  (mirrors the template sheet layout)
# ═══════════════════════════════════════════════════════════════════════════

def generate_xlsx(
    snap: UsageSnapshot,
    cx: CoralogixSizing,
    all_pairs: list[tuple["UsageSnapshot", "CoralogixSizing"]] | None = None,
) -> bytes | None:
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()

    # ── Helper styles ─────────────────────────────────────────────────────
    _GREEN   = "008F61"
    _GREEN_L = "00B37A"
    _INK     = "1A2332"
    _LGRAY   = "F5F5F5"
    _DGRAY   = "3C3C3C"
    _WHITE   = "FFFFFF"

    def hdr_cell(ws, row, col, value, bg=_GREEN, fg=_WHITE, bold=True, sz=11):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def label_cell(ws, row, col, value, bold=False, bg=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=10)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    def val_cell(ws, row, col, value, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal="right")
        if number_format:
            c.number_format = number_format
        return c

    # ════════════════════════════════════════════════════════════════════
    # Sheet 1 — Bill Overview
    # ════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Bill Overview"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22
    ws1.row_dimensions[1].height = 28

    hdr_cell(ws1, 1, 1, f"Datadog Bill Overview — {snap.month}", bg=_GREEN, sz=13)
    ws1.merge_cells("A1:D1")
    ws1.cell(row=2, column=1, value=f"Account: {snap.account_name or 'N/A'}  |  Site: {snap.site}  |  Pulled: {snap.pulled_at[:10]}")
    ws1.merge_cells("A2:D2")
    ws1.cell(row=3, column=1, value=f"Note: {FRESHNESS_NOTE}").font = Font(italic=True, size=9, color="666666")
    ws1.merge_cells("A3:D3")

    hdr_cell(ws1, 5, 1, "Product Metric",   bg=_DGRAY, sz=10)
    hdr_cell(ws1, 5, 2, "Raw Value",        bg=_DGRAY, sz=10)
    hdr_cell(ws1, 5, 3, "Formatted",        bg=_DGRAY, sz=10)
    hdr_cell(ws1, 5, 4, "Unit",             bg=_DGRAY, sz=10)

    tiles = [
        ("Infrastructure",           None,                          None,     ""),
        ("Infra Hosts",               snap.infra_hosts,              _fmt(snap.infra_hosts, 0), "hosts (avg concurrent)"),
        ("APM Hosts",                snap.apm_hosts,                _fmt(snap.apm_hosts, 0),   "hosts (avg concurrent)"),
        ("Container Hours",          snap.container_hours,          _fmt(snap.container_hours), "container-hours/month"),
        ("Containers (avg concurrent)", snap.containers,            _fmt(snap.containers, 0),  "containers"),
        ("Network Hosts",            snap.network_hosts,            _fmt(snap.network_hosts, 0), "hosts"),
        ("DBM Hosts",                snap.dbm_hosts,                _fmt(snap.dbm_hosts, 0),   "hosts"),
        ("",                         None,                          None,     ""),
        ("Custom Metrics",           None,                          None,     ""),
        ("Custom Metrics",           snap.custom_metrics,           _fmt(snap.custom_metrics), "time series"),
        ("Ingested Custom Metrics",  snap.ingested_custom_metrics,  _fmt(snap.ingested_custom_metrics), "time series"),
        ("",                         None,                          None,     ""),
        ("Logs",                     None,                          None,     ""),
        ("Ingested Logs",            snap.ingested_logs_bytes,      _bytes_to_tb(snap.ingested_logs_bytes), "bytes"),
        ("Indexed Logs (3 Day)",     snap.indexed_logs_3day,        _fmt(snap.indexed_logs_3day), "events"),
        ("Indexed Logs (7 Day)",     snap.indexed_logs_7day,        _fmt(snap.indexed_logs_7day), "events"),
        ("Indexed Logs (15 Day)",    snap.indexed_logs_15day,       _fmt(snap.indexed_logs_15day), "events"),
        ("Indexed Logs (30 Day)",    snap.indexed_logs_30day,       _fmt(snap.indexed_logs_30day), "events"),
        ("Indexed Logs (45 Day)",    snap.indexed_logs_45day,       _fmt(snap.indexed_logs_45day), "events"),
        ("Indexed Logs (60 Day)",    snap.indexed_logs_60day,       _fmt(snap.indexed_logs_60day), "events"),
        ("Indexed Logs (90 Day)",    snap.indexed_logs_90day,       _fmt(snap.indexed_logs_90day), "events"),
        ("Indexed Logs (180 Day)",   snap.indexed_logs_180day,      _fmt(snap.indexed_logs_180day), "events"),
        ("Indexed Logs (360 Day)",   snap.indexed_logs_360day,      _fmt(snap.indexed_logs_360day), "events"),
        ("Indexed Logs (Live Search)", snap.indexed_logs_live,      _fmt(snap.indexed_logs_live), "events"),
        ("Indexed Logs (Rehydrated)", snap.indexed_logs_rehydrated, _fmt(snap.indexed_logs_rehydrated), "events"),
        ("SIEM / Security Logs",     snap.security_logs_bytes,      _bytes_to_gb(snap.security_logs_bytes), "bytes"),
        ("",                         None,                          None,     ""),
        ("APM / Tracing",            None,                          None,     ""),
        ("Ingested Spans",           snap.ingested_spans_bytes,     _bytes_to_tb(snap.ingested_spans_bytes), "bytes"),
        ("Indexed Spans",            snap.indexed_spans,            _fmt(snap.indexed_spans), "events"),
        ("Profiled Hosts",           snap.profiled_hosts,           _fmt(snap.profiled_hosts, 0), "hosts"),
        ("Profiled Container Hours", snap.profiled_containers,      _fmt(snap.profiled_containers), "container-hours"),
        ("APM Fargate Tasks",        snap.apm_fargate,              _fmt(snap.apm_fargate, 0), "tasks"),
        ("Profiled Fargate Tasks",   snap.profiled_fargate,         _fmt(snap.profiled_fargate, 0), "tasks"),
        ("Custom Events",            snap.custom_events,            _fmt(snap.custom_events), "events"),
        ("",                         None,                          None,     ""),
        ("Serverless",               None,                          None,     ""),
        ("Serverless Workload Functions", snap.serverless_functions, _fmt(snap.serverless_functions, 0), "functions"),
        ("Serverless Invocations",   snap.serverless_invocations,   _fmt(snap.serverless_invocations), "invocations/month"),
        ("Serverless App Instances", snap.serverless_app_instances, _fmt(snap.serverless_app_instances, 0), "instances"),
        ("Fargate Tasks",            snap.fargate_tasks,            _fmt(snap.fargate_tasks, 0), "tasks"),
        ("",                         None,                          None,     ""),
        ("RUM",                      None,                          None,     ""),
        ("RUM Investigate",          snap.rum_sessions,             _fmt(snap.rum_sessions), "sessions/month"),
        ("RUM Measure",              snap.rum_lite_sessions,        _fmt(snap.rum_lite_sessions), "sessions/month"),
        ("Session Replay",           snap.rum_replay,               _fmt(snap.rum_replay), "sessions/month"),
        ("Error Tracking Events",    snap.rum_errors,               _fmt(snap.rum_errors), "events/month"),
        ("",                         None,                          None,     ""),
        ("Synthetics",               None,                          None,     ""),
        ("Synthetics API Test Runs",     snap.synthetics_api,       _fmt(snap.synthetics_api), "test runs/month"),
        ("Synthetics Browser Test Runs", snap.synthetics_browser,   _fmt(snap.synthetics_browser), "test runs/month"),
        ("",                         None,                          None,     ""),
        ("Other",                    None,                          None,     ""),
        ("Incident Management Seats",    snap.incident_management_seats,   _fmt(snap.incident_management_seats, 0), "seats"),
        ("Test Optimization Committers", snap.test_optimization_committers, _fmt(snap.test_optimization_committers, 0), "committers"),
        ("Test Optimization Spans",      snap.test_optimization_spans,     _fmt(snap.test_optimization_spans), "spans"),
        ("Product Analytics Sessions",   snap.product_analytics_sessions,  _fmt(snap.product_analytics_sessions), "sessions/month"),
        ("App Builder Published Apps",   snap.app_builder_apps,            _fmt(snap.app_builder_apps, 0), "apps"),
        ("Bits AI SRE Investigations",   snap.bits_ai_investigations,      _fmt(snap.bits_ai_investigations, 0), "investigations/month"),
    ]

    section_headers = {"Infrastructure", "Custom Metrics", "Logs", "APM / Tracing",
                       "Serverless", "RUM", "Synthetics", "Other"}

    for i, (name, raw, formatted, unit) in enumerate(tiles, start=6):
        row = i
        if name in section_headers:
            c = ws1.cell(row=row, column=1, value=name)
            c.font = Font(bold=True, size=10, color=_WHITE)
            c.fill = PatternFill("solid", fgColor=_GREEN_L)
            ws1.merge_cells(f"A{row}:D{row}")
        elif name == "":
            pass
        else:
            fill = PatternFill("solid", fgColor=_LGRAY) if row % 2 == 0 else None
            c = ws1.cell(row=row, column=1, value=name)
            c.font = Font(size=10)
            if fill:
                c.fill = fill

            c2 = ws1.cell(row=row, column=2, value=raw if raw is not None else "")
            c2.font = Font(size=10)
            c2.alignment = Alignment(horizontal="right")
            if fill:
                c2.fill = fill

            c3 = ws1.cell(row=row, column=3, value=formatted or "")
            c3.font = Font(size=10, bold=True)
            c3.alignment = Alignment(horizontal="right")
            if fill:
                c3.fill = fill

            c4 = ws1.cell(row=row, column=4, value=unit)
            c4.font = Font(size=10, color="666666")
            if fill:
                c4.fill = fill

    # ════════════════════════════════════════════════════════════════════
    # Sheet 2 — Coralogix Sizing  (mirrors the Excel template layout)
    # ════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Coralogix Sizing")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 36
    ws2.row_dimensions[1].height = 28

    hdr_cell(ws2, 1, 1, f"Coralogix Sizing Estimate — {snap.month}", bg=_GREEN_L, sz=13)
    ws2.merge_cells("A1:D1")
    ws2.cell(row=2, column=1, value=f"Based on Datadog usage for {snap.account_name or 'account'}. "
             "All assumptions are listed below and can be adjusted.")
    ws2.merge_cells("A2:D2")

    # ── Assumptions block ────────────────────────────────────────────────
    hdr_cell(ws2, 4, 1, "Sizing Assumptions", bg=_DGRAY, sz=10)
    ws2.merge_cells("A4:D4")
    assumptions = [
        ("Avg log line size", AVG_LOG_SIZE_KB, "KB"),
        ("Avg span size",     AVG_SPAN_SIZE_KB, "KB"),
        ("Time series per host/container", TS_PER_HOST, "TS"),
        ("TS-to-Units factor (metrics)",   TS_TO_UNITS, "Units/TS"),
        ("Days per month",    DAYS_PER_MONTH, "days"),
        ("Log tier split",    f"{int(LOG_TIER_MON*100)}% Mon / {int(LOG_TIER_COMP*100)}% Comp", ""),
        ("Span tier split",   f"{int(SPAN_TIER_MON*100)}% Mon / {int(SPAN_TIER_COMP*100)}% Comp", ""),
    ]
    for i, (label, val, unit) in enumerate(assumptions, start=5):
        ws2.cell(row=i, column=1, value=label).font = Font(size=10)
        ws2.cell(row=i, column=2, value=val).font = Font(size=10, bold=True)
        ws2.cell(row=i, column=3, value=unit).font = Font(size=10, color="666666")

    # ── DD Usage Input block (mirrors template rows 1-10) ────────────────
    hdr_cell(ws2, 12, 1, "DD Usage Inputs", bg=_DGRAY, sz=10)
    ws2.merge_cells("A12:D12")

    dd_inputs = [
        ("Infra Hosts",                         snap.infra_hosts,            "hosts (avg concurrent)"),
        ("APM Hosts",                           snap.apm_hosts,              "hosts (avg concurrent)"),
        ("Profiled Hosts",                      snap.profiled_hosts,         "hosts (avg concurrent)"),
        ("Containers",                          snap.containers,             "containers (avg concurrent)"),
        ("Custom Metrics",                      snap.custom_metrics,         "time series"),
        ("Ingested Logs",                       snap.ingested_logs_bytes/1e9,"GB/month"),
        ("Security Logs",                       snap.security_logs_bytes/1e9,"GB/month"),
        ("Indexed Logs — Live & Rehydrated",    (snap.indexed_logs_3day+snap.indexed_logs_7day+snap.indexed_logs_live+snap.indexed_logs_rehydrated)/1e6, "Million events"),
        ("Indexed Logs — 15 Day",               snap.indexed_logs_15day/1e6, "Million events"),
        ("Indexed Logs — 90 Day (30-360d sum)", (snap.indexed_logs_30day+snap.indexed_logs_45day+snap.indexed_logs_60day+snap.indexed_logs_90day+snap.indexed_logs_180day+snap.indexed_logs_360day)/1e6, "Million events"),
        ("Ingested Spans",                      snap.ingested_spans_bytes/1e9,"GB/month"),
        ("Indexed Spans",                       snap.indexed_spans/1e6,       "Million events"),
        ("Custom Events",                       snap.custom_events,           "events/month"),
        ("Serverless Functions (monthly avg)",  snap.serverless_functions,    "functions"),
        ("Serverless Invocations (monthly)",    snap.serverless_invocations,  "invocations"),
        ("Fargate Tasks",                       snap.fargate_tasks,           "tasks"),
        ("APM Fargate Tasks",                   snap.apm_fargate,             "tasks"),
        ("Profiled Fargate Tasks",              snap.profiled_fargate,        "tasks"),
        ("RUM Sessions",                        snap.rum_sessions,            "sessions/month"),
        ("Error Tracking Events",               snap.rum_errors,              "events/month"),
    ]
    hdr_cell(ws2, 13, 1, "Input Metric", bg=_DGRAY, sz=10)
    hdr_cell(ws2, 13, 2, "Value",        bg=_DGRAY, sz=10)
    hdr_cell(ws2, 13, 3, "Unit",         bg=_DGRAY, sz=10)
    for i, (label, val, unit) in enumerate(dd_inputs, start=14):
        fill = PatternFill("solid", fgColor=_LGRAY) if i % 2 == 0 else None
        c1 = ws2.cell(row=i, column=1, value=label)
        c1.font = Font(size=10)
        if fill: c1.fill = fill
        c2 = ws2.cell(row=i, column=2, value=round(val, 4) if isinstance(val, float) else val)
        c2.font = Font(size=10, bold=True)
        c2.alignment = Alignment(horizontal="right")
        if fill: c2.fill = fill
        c3 = ws2.cell(row=i, column=3, value=unit)
        c3.font = Font(size=10, color="666666")
        if fill: c3.fill = fill

    # ── Coralogix Sizing Output ───────────────────────────────────────────
    out_row = 14 + len(dd_inputs) + 2

    sections = [
        ("LOGS", _GREEN_L, [
            ("Total Ingested Logs",           cx.total_ingested_logs_gb_month, "GB/month",   "ingested_logs_gb + security_logs_gb"),
            ("Total Indexed Logs (events)",   cx.total_indexed_logs_count,     "events/month","all retention tiers summed"),
            ("Total Indexed Logs Size",       cx.indexed_logs_size_gb_month,   "GB/month",   "count × avg_log_size_kb × 1024 ÷ 1024³"),
            ("Indexed Percentage",            cx.indexed_pct_logs * 100,       "%",          "indexed_size_gb / ingested_gb"),
            ("Daily Ingested Logs",           cx.daily_logs_gb,                "GB/day",     "ingested_gb ÷ 30"),
            ("  Monitoring (70%)",            cx.daily_logs_mon_gb,            "GB/day",     "daily × 0.70"),
            ("  Compliance (30%)",            cx.daily_logs_comp_gb,           "GB/day",     "daily × 0.30"),
        ]),
        ("METRICS", _GREEN_L, [
            ("Host Count (all types incl. fargate)", cx.host_count,         "hosts",       "infra+apm+profiled+network+fargate"),
            ("Container Count",               cx.container_count,              "containers", "containers + profiled_containers"),
            ("Host+Container TS",             cx.host_container_ts,            "NumSeries",  "(hosts+containers) × 750"),
            ("Serverless Functions TS",       cx.sw_func_ts,                   "NumSeries",  "daily_functions × 0.30"),
            ("Serverless Invocations TS",     cx.sw_invoc_ts,                  "NumSeries",  "daily_invocations × 0.30"),
            ("Total TimeSeries (NumSeries)",  cx.total_ts,                     "NumSeries",  "host_cont_ts + custom_metrics + sw_ts"),
            ("Metrics Units / Day",           cx.metrics_units_per_day,        "Units/day",  "total_ts × 3.3e-5"),
        ]),
        ("TRACING", _GREEN_L, [
            ("Ingested Spans",                cx.ingested_spans_gb_month,      "GB/month",   "twol_ingested_events_bytes_sum ÷ 1e9"),
            ("Indexed Spans",                 cx.indexed_spans_gb_month,       "GB/month",   "(indexed_spans+custom_events) × 1.5 KB ÷ 1024²"),
            ("Indexed Span Percentage",       cx.indexed_pct_spans * 100,      "%",          "indexed_gb / ingested_gb"),
            ("Daily Ingested Spans",          cx.daily_spans_ingest_gb,        "GB/day",     "ingested_gb ÷ 30"),
            ("  Monitoring (10%)",            cx.daily_spans_mon_gb,           "GB/day",     f"daily × {SPAN_TIER_MON}"),
            ("  Compliance (90%)",            cx.daily_spans_comp_gb,          "GB/day",     f"daily × {SPAN_TIER_COMP}"),
            ("  Indexed (GB/day)",            cx.daily_spans_indexed_gb,       "GB/day",     "daily × indexed_pct"),
            ("  Archive (GB/day)",            cx.daily_spans_archive_gb,       "GB/day",     "daily - indexed_daily"),
        ]),
        ("RUM", _GREEN_L, [
            ("RUM Sessions / Month",              cx.rum_sessions_monthly,         "sessions",   "rum_total_session_count"),
            ("Total Sessions / Day",              cx.rum_sessions_daily,           "sessions/day", "monthly ÷ 30"),
            ("Session Recording / Day (Replay)",  cx.rum_session_recording_daily,  "sessions/day", "replay ÷ 30"),
            ("RUM Errors / Day",              cx.rum_errors_per_day,           "events/day", "error_tracking_events ÷ 30"),
        ]),
        ("SUMMARY — Apply to Coralogix TCO Calculator", _GREEN, [
            ("Logs GB/day",                   cx.daily_logs_gb,                "GB/day",     ""),
            ("Metrics NumSeries",             cx.total_ts,                     "NumSeries",  ""),
            ("Tracing GB/day",                cx.daily_spans_ingest_gb,        "GB/day",     ""),
            ("RUM Total Sessions/day",        cx.rum_sessions_daily,           "sessions/day", ""),
            ("RUM Session Recording/day",     cx.rum_session_recording_daily,  "sessions/day", "replay"),
        ]),
    ]

    for section_name, section_color, rows in sections:
        hdr_cell(ws2, out_row, 1, section_name, bg=section_color, sz=10)
        ws2.merge_cells(f"A{out_row}:D{out_row}")
        out_row += 1
        hdr_cell(ws2, out_row, 1, "Metric",     bg=_DGRAY, sz=9)
        hdr_cell(ws2, out_row, 2, "Value",      bg=_DGRAY, sz=9)
        hdr_cell(ws2, out_row, 3, "Unit",       bg=_DGRAY, sz=9)
        hdr_cell(ws2, out_row, 4, "Formula",    bg=_DGRAY, sz=9)
        out_row += 1
        for i, (label, val, unit, formula) in enumerate(rows):
            fill = PatternFill("solid", fgColor=_LGRAY) if i % 2 == 0 else None
            c1 = ws2.cell(row=out_row, column=1, value=label)
            c1.font = Font(size=10)
            if fill: c1.fill = fill
            c2 = ws2.cell(row=out_row, column=2,
                          value=round(val, 4) if isinstance(val, float) else val)
            c2.font = Font(size=10, bold=True)
            c2.alignment = Alignment(horizontal="right")
            if fill: c2.fill = fill
            c3 = ws2.cell(row=out_row, column=3, value=unit)
            c3.font = Font(size=10, color="666666")
            if fill: c3.fill = fill
            c4 = ws2.cell(row=out_row, column=4, value=formula)
            c4.font = Font(size=9, color="888888", italic=True)
            if fill: c4.fill = fill
            out_row += 1
        out_row += 1

    # ── Trend sheet (when multiple months available) ──────────────────────
    if all_pairs and len(all_pairs) > 1:
        trends = compute_trends(all_pairs)
        ws4 = wb.create_sheet("Trend Analysis")
        ws4.column_dimensions["A"].width = 30
        for col_i in range(len(trends)):
            ws4.column_dimensions[chr(66 + col_i * 2)].width = 16
            ws4.column_dimensions[chr(67 + col_i * 2)].width = 12

        # Header
        hdr_cell(ws4, 1, 1, "Trend Analysis — TCO Metrics", bg=_GREEN, sz=13)
        ws4.merge_cells(f"A1:{chr(65 + len(trends)*2)}1")

        # Month headers
        ws4.cell(row=2, column=1, value="Metric").font = Font(bold=True, size=10)
        for i, t in enumerate(trends):
            c = ws4.cell(row=2, column=2 + i * 2, value=t["month"])
            c.font = Font(bold=True, color=_WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=_GREEN_L)
            c.alignment = Alignment(horizontal="center")
            if i > 0:
                mc = ws4.cell(row=2, column=3 + i * 2 - 2 + 1, value="MoM %")
                mc.font = Font(bold=True, color=_WHITE, size=9)
                mc.fill = PatternFill("solid", fgColor="888888")
                mc.alignment = Alignment(horizontal="center")

        metrics_def = [
            ("Logs",              "GB/day",       "logs_gb_day",    "logs_pct"),
            ("Metrics",           "NumSeries",    "metrics_ts",     "metrics_pct"),
            ("Tracing",           "GB/day",       "tracing_gb_day", "tracing_pct"),
            ("RUM Sessions",      "sessions/day", "rum_day",        "rum_pct"),
            ("RUM Recordings",    "sessions/day", "rum_rec_day",    None),
        ]
        for row_i, (label, unit, val_key, pct_key) in enumerate(metrics_def):
            r = 3 + row_i
            fill = PatternFill("solid", fgColor="F9F7FC") if row_i % 2 == 0 else None
            c = ws4.cell(row=r, column=1, value=f"{label} ({unit})")
            c.font = Font(size=10, bold=True)
            if fill: c.fill = fill
            for i, t in enumerate(trends):
                val = t.get(val_key, 0) or 0
                vc = ws4.cell(row=r, column=2 + i * 2, value=round(val, 2))
                vc.font = Font(size=10)
                vc.alignment = Alignment(horizontal="right")
                if fill: vc.fill = fill
                if i > 0 and pct_key:
                    pct = t.get(pct_key)
                    if pct is not None:
                        pc = ws4.cell(row=r, column=3 + i * 2 - 2 + 1, value=round(pct, 1))
                        pc.font = Font(size=10, bold=True,
                                       color="E04B2A" if pct > 5 else ("22A06B" if pct < -5 else "888888"))
                        pc.alignment = Alignment(horizontal="right")
                        if fill: pc.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 9 — HTML report  (self-contained, no external dependencies)
# ═══════════════════════════════════════════════════════════════════════════

def generate_html(
    snap: UsageSnapshot,
    cx: CoralogixSizing,
    all_pairs: list[tuple["UsageSnapshot", "CoralogixSizing"]] | None = None,
    trends: list[dict] | None = None,
) -> bytes:
    """Light-theme HTML report using Coralogix green brand colours."""

    # Coralogix brand: green primary, not orange
    CX_GREEN   = "#00B37A"
    CX_GREEN_D = "#008F61"
    CX_INK     = "#1A2332"
    CX_MUTED   = "#5B6B7C"
    CX_LINE    = "#E3E8EE"
    CX_SOFT    = "#F4F7F5"
    CX_WHITE   = "#FFFFFF"
    CX_GROW    = "#C0392B"   # growth = attention (red)
    CX_DROP    = "#008F61"   # decline indicator (green)
    CX_FLAT    = "#7A8694"

    def _fv(val: float, dec: int = 1) -> str:
        if not val:
            return "—"
        return f"{val:,.{dec}f}" if dec else f"{val:,.0f}"

    def _drow(label: str, val: str, note: str = "") -> str:
        return (f'<tr><td>{label}</td><td class="r">{val}</td>'
                f'<td class="note">{note}</td></tr>')

    def _tile_row(items: list[tuple[str, str, str]]) -> str:
        out = ""
        for label, val, sub in items:
            if val in ("0", "0.0", "—", "0.00"):
                continue
            out += (
                f'<div class="src">'
                f'<div class="src-l">{label}</div>'
                f'<div class="src-v">{val}</div>'
                f'{f"<div class=src-s>{sub}</div>" if sub else ""}'
                f'</div>'
            )
        return out

    # ── Trend section ─────────────────────────────────────────────────────
    trend_metrics = [
        ("Logs",           "GB/day",    "logs_gb_day",    "logs_pct",    1),
        ("Metrics",        "NumSeries", "metrics_ts",     "metrics_pct", 0),
        ("Tracing",        "GB/day",    "tracing_gb_day", "tracing_pct", 1),
        ("RUM Sessions",   "sess/day",  "rum_day",        "rum_pct",     0),
        ("RUM Recordings", "rec/day",   "rum_rec_day",    None,          0),
    ]

    def _classify(pct_key: str | None):
        if not trends or not pct_key:
            return ("—", CX_FLAT)
        pcts = [t[pct_key] for t in trends[1:] if t.get(pct_key) is not None]
        if not pcts:
            return ("—", CX_FLAT)
        avg = sum(pcts) / len(pcts)
        if avg > 10:  return ("Fast growth", CX_GROW)
        if avg > 2:   return ("Growing",     "#D97706")
        if avg < -10: return ("Fast decline", CX_DROP)
        if avg < -2:  return ("Declining",   CX_DROP)
        return ("Stable", CX_FLAT)

    trend_html = ""
    if trends and len(trends) > 1:
        rows = ""
        for label, unit, vk, pk, dec in trend_metrics:
            cls_txt, cls_col = _classify(pk)
            cells = ""
            for i, t in enumerate(trends):
                v = _fv(t.get(vk, 0), dec)
                pct = t.get(pk) if (pk and i > 0) else None
                if pct is None:
                    pct_s = ""
                elif pct > 5:
                    pct_s = f'<span class="chg up">{pct:+.1f}%</span>'
                elif pct < -5:
                    pct_s = f'<span class="chg dn">{pct:+.1f}%</span>'
                else:
                    pct_s = f'<span class="chg flat">{pct:+.1f}%</span>'
                cells += f'<td class="r">{v}{pct_s}</td>'
            badge = (
                f'<span class="badge" style="color:{cls_col};border-color:{cls_col}">'
                f'{cls_txt}</span>'
            ) if pk else ""
            rows += (
                f'<tr><td class="m">{label}'
                f'<span class="u">{unit}</span>{badge}</td>{cells}</tr>'
            )
        month_ths = "".join(f"<th>{t['month']}</th>" for t in trends)
        trend_html = f"""
<section class="panel">
  <div class="panel-h">
    <h2>Growth trend</h2>
    <span class="panel-sub">{trends[0]['month']} → {trends[-1]['month']}</span>
  </div>
  <table class="trend">
    <thead><tr><th>Metric</th>{month_ths}</tr></thead>
    <tbody>{rows}</tbody>
  </table>
  <p class="legend">% under each value = change vs prior month · red = growth · green = decline · grey = stable (±5%)</p>
</section>"""

    # ── Source tiles ──────────────────────────────────────────────────────
    infra = _tile_row([
        ("Infra Hosts",     _fmt(snap.infra_hosts, 0),    "avg concurrent"),
        ("APM Hosts",       _fmt(snap.apm_hosts, 0),      "avg concurrent"),
        ("Containers",      _fmt(snap.containers, 0),     "avg concurrent"),
        ("Profiled Hosts",  _fmt(snap.profiled_hosts, 0), ""),
        ("Fargate Tasks",   _fmt(snap.fargate_tasks, 0),  ""),
        ("Network Hosts",   _fmt(snap.network_hosts, 0),  ""),
        ("DBM Hosts",       _fmt(snap.dbm_hosts, 0),      ""),
        ("Custom Metrics",  _fmt(snap.custom_metrics),    ""),
    ])
    logs = _tile_row([
        ("Ingested Logs",   _bytes_to_tb(snap.ingested_logs_bytes), ""),
        ("Indexed 3d",      _fmt(snap.indexed_logs_3day),  ""),
        ("Indexed 7d",      _fmt(snap.indexed_logs_7day),  ""),
        ("Indexed 15d",     _fmt(snap.indexed_logs_15day), ""),
        ("Indexed 30d",     _fmt(snap.indexed_logs_30day), ""),
        ("Indexed 90d",     _fmt(snap.indexed_logs_90day), ""),
        ("SIEM",            _bytes_to_gb(snap.security_logs_bytes), ""),
    ])
    apm = _tile_row([
        ("Ingested Spans",  _bytes_to_tb(snap.ingested_spans_bytes), ""),
        ("Indexed Spans",   _fmt(snap.indexed_spans), ""),
    ])
    rum = _tile_row([
        ("Sessions",        _fmt(snap.rum_sessions), ""),
        ("Session Replay",  _fmt(snap.rum_replay),   ""),
        ("Errors",          _fmt(snap.rum_errors),   ""),
    ])
    synth = _tile_row([
        ("API Test Runs",     _fmt(snap.synthetics_api, 0),     "monthly"),
        ("Browser Test Runs", _fmt(snap.synthetics_browser, 0), "monthly"),
    ])

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Datadog → Coralogix — {snap.month}</title>
<style>
/* DM Sans loaded from system/fallback stack — no external font request */
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',system-ui,sans-serif;background:{CX_SOFT};color:{CX_INK};font-size:14px;line-height:1.5}}
a{{color:{CX_GREEN_D}}}

.top{{background:{CX_WHITE};border-bottom:1px solid {CX_LINE};padding:18px 32px;display:flex;align-items:center;gap:14px}}
.mark{{width:10px;height:10px;border-radius:50%;background:{CX_GREEN};box-shadow:0 0 0 4px rgba(0,179,122,.18)}}
.top h1{{font-size:17px;font-weight:800;letter-spacing:-.02em}}
.top .meta{{font-size:12px;color:{CX_MUTED};margin-top:2px}}
.top .meta b{{color:{CX_INK};font-weight:600}}

.wrap{{max-width:1080px;margin:0 auto;padding:28px 32px 48px}}

.action{{background:{CX_WHITE};border:1px solid {CX_LINE};border-radius:14px;padding:22px 24px;box-shadow:0 1px 2px rgba(26,35,50,.04)}}
.action-h{{display:flex;align-items:baseline;justify-content:space-between;gap:12px;margin-bottom:16px;flex-wrap:wrap}}
.action-h h2{{font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.06em;color:{CX_GREEN_D}}}
.action-h p{{font-size:12px;color:{CX_MUTED}}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px}}
@media(max-width:900px){{.kpis{{grid-template-columns:repeat(2,minmax(0,1fr))}}}}
.kpi{{background:{CX_SOFT};border:1px solid {CX_LINE};border-radius:10px;padding:14px 14px 12px}}
.kpi .l{{font-size:11px;font-weight:700;color:{CX_GREEN_D};text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px}}
.kpi .v{{font-size:22px;font-weight:800;letter-spacing:-.03em;color:{CX_INK};line-height:1.1}}
.kpi .u{{font-size:11px;color:{CX_MUTED};margin-top:4px}}
.kpi .d{{font-size:11px;color:{CX_MUTED};margin-top:8px;line-height:1.45}}

.panel{{background:{CX_WHITE};border:1px solid {CX_LINE};border-radius:14px;padding:20px 22px;margin-top:18px}}
.panel-h{{display:flex;align-items:baseline;gap:10px;margin-bottom:12px}}
.panel-h h2{{font-size:15px;font-weight:800;color:{CX_INK}}}
.panel-sub{{font-size:12px;color:{CX_MUTED}}}

table{{width:100%;border-collapse:collapse}}
th{{text-align:left;font-size:11px;font-weight:700;color:{CX_MUTED};text-transform:uppercase;letter-spacing:.04em;padding:8px 10px;border-bottom:1px solid {CX_LINE}}}
th:not(:first-child),.r{{text-align:right}}
td{{padding:10px;border-bottom:1px solid {CX_LINE};vertical-align:top}}
tr:last-child td{{border-bottom:none}}
.m{{font-weight:700;color:{CX_INK}}}
.u{{display:block;font-size:11px;color:{CX_MUTED};font-weight:500;margin-top:1px}}
.badge{{display:inline-block;margin-top:6px;font-size:10px;font-weight:700;padding:2px 8px;border:1px solid;border-radius:999px}}
.chg{{display:block;font-size:11px;font-weight:700;margin-top:3px}}
.chg.up{{color:{CX_GROW}}}
.chg.dn{{color:{CX_DROP}}}
.chg.flat{{color:{CX_FLAT}}}
.legend{{font-size:11px;color:{CX_MUTED};margin-top:10px}}

.srcs{{display:flex;flex-wrap:wrap;gap:8px}}
.src{{background:{CX_SOFT};border:1px solid {CX_LINE};border-radius:8px;padding:10px 12px;min-width:118px;flex:1 1 118px;max-width:160px}}
.src-l{{font-size:11px;color:{CX_MUTED};font-weight:600}}
.src-v{{font-size:15px;font-weight:800;color:{CX_INK};margin-top:2px}}
.src-s{{font-size:10px;color:{CX_MUTED};margin-top:2px}}

.detail .r{{color:{CX_GREEN_D};font-weight:800;font-variant-numeric:tabular-nums}}
.note{{font-size:11px;color:{CX_MUTED}}}

.assumptions{{margin-top:18px;font-size:12px;color:{CX_MUTED};line-height:1.7}}
.assumptions b{{color:{CX_INK}}}
.ftr{{text-align:center;padding:18px;font-size:11px;color:{CX_MUTED}}}
</style>
</head>
<body>

<header class="top">
  <div class="mark"></div>
  <div>
    <h1>Datadog → Coralogix sizing report</h1>
    <div class="meta">{snap.month} · <b>{snap.account_name or "N/A"}</b> · {snap.site} · generated {snap.pulled_at[:10]}</div>
  </div>
</header>

<main class="wrap">

<section class="action">
  <div class="action-h">
    <h2>Enter these in the Coralogix TCO sheet</h2>
    <p>Primary sizing inputs for {snap.month}</p>
  </div>
  <div class="kpis">
    <div class="kpi">
      <div class="l">Logs</div>
      <div class="v">{cx.daily_logs_gb:,.1f}</div>
      <div class="u">GB / day</div>
      <div class="d">Mon {int(LOG_TIER_MON*100)}%: {cx.daily_logs_mon_gb:,.1f}<br>Comp {int(LOG_TIER_COMP*100)}%: {cx.daily_logs_comp_gb:,.1f}</div>
    </div>
    <div class="kpi">
      <div class="l">Metrics</div>
      <div class="v">{cx.total_ts:,.0f}</div>
      <div class="u">NumSeries</div>
    </div>
    <div class="kpi">
      <div class="l">Tracing</div>
      <div class="v">{cx.daily_spans_ingest_gb:,.1f}</div>
      <div class="u">GB / day</div>
      <div class="d">Mon {int(SPAN_TIER_MON*100)}%: {cx.daily_spans_mon_gb:,.1f}<br>Comp {int(SPAN_TIER_COMP*100)}%: {cx.daily_spans_comp_gb:,.1f}</div>
    </div>
    <div class="kpi">
      <div class="l">RUM Sessions</div>
      <div class="v">{cx.rum_sessions_daily:,.0f}</div>
      <div class="u">sessions / day</div>
    </div>
    <div class="kpi">
      <div class="l">RUM Recording</div>
      <div class="v">{cx.rum_session_recording_daily:,.0f}</div>
      <div class="u">recordings / day</div>
    </div>
  </div>
</section>

<section class="action" style="margin-top:14px">
  <div class="action-h">
    <h2>Enter these in Checkly</h2>
    <p>Datadog Synthetics → Checkly sizing for {snap.month}</p>
  </div>
  <div class="kpis">
    <div class="kpi">
      <div class="l">API checks</div>
      <div class="v">{cx.synthetics_api_daily:,.0f}</div>
      <div class="u">test runs / day</div>
      <div class="d">{cx.synthetics_api_monthly:,.0f} / month</div>
    </div>
    <div class="kpi">
      <div class="l">Browser checks</div>
      <div class="v">{cx.synthetics_browser_daily:,.0f}</div>
      <div class="u">test runs / day</div>
      <div class="d">{cx.synthetics_browser_monthly:,.0f} / month</div>
    </div>
  </div>
</section>

{trend_html}

<section class="panel">
  <div class="panel-h"><h2>Datadog source — infra &amp; metrics</h2><span class="panel-sub">{snap.month}</span></div>
  <div class="srcs">{infra}</div>
</section>
<section class="panel">
  <div class="panel-h"><h2>Datadog source — logs</h2></div>
  <div class="srcs">{logs}</div>
</section>
<section class="panel">
  <div class="panel-h"><h2>Datadog source — tracing</h2></div>
  <div class="srcs">{apm}</div>
</section>
<section class="panel">
  <div class="panel-h"><h2>Datadog source — RUM</h2></div>
  <div class="srcs">{rum}</div>
</section>
<section class="panel">
  <div class="panel-h"><h2>Datadog source — Synthetics (→ Checkly)</h2></div>
  <div class="srcs">{synth}</div>
</section>

<section class="panel">
  <div class="panel-h"><h2>Sizing detail — Synthetics → Checkly</h2></div>
  <table class="detail">
    <thead><tr><th>Metric</th><th>Value</th><th>Note</th></tr></thead>
    <tbody>
      {_drow("API checks / day",     f"{cx.synthetics_api_daily:,.0f}",     "÷ 30")}
      {_drow("API checks / month",   f"{cx.synthetics_api_monthly:,.0f}")}
      {_drow("Browser checks / day", f"{cx.synthetics_browser_daily:,.0f}", "÷ 30")}
      {_drow("Browser checks / month", f"{cx.synthetics_browser_monthly:,.0f}")}
    </tbody>
  </table>
</section>

<section class="panel">
  <div class="panel-h"><h2>Sizing detail — logs</h2></div>
  <table class="detail">
    <thead><tr><th>Metric</th><th>Value</th><th>Note</th></tr></thead>
    <tbody>
      {_drow("Total ingested", f"{cx.total_ingested_logs_gb_month:,.1f} GB/mo")}
      {_drow("Total indexed",  f"{cx.total_indexed_logs_count:,.0f} events", "all retention tiers")}
      {_drow("Daily ingested", f"{cx.daily_logs_gb:,.1f} GB/day", "÷ 30")}
      {_drow("→ Monitoring",   f"{cx.daily_logs_mon_gb:,.1f} GB/day", f"{int(LOG_TIER_MON*100)}%")}
      {_drow("→ Compliance",   f"{cx.daily_logs_comp_gb:,.1f} GB/day", f"{int(LOG_TIER_COMP*100)}%")}
    </tbody>
  </table>
</section>
<section class="panel">
  <div class="panel-h"><h2>Sizing detail — metrics</h2></div>
  <table class="detail">
    <thead><tr><th>Metric</th><th>Value</th><th>Note</th></tr></thead>
    <tbody>
      {_drow("Hosts (all types)", f"{cx.host_count:,.0f}", "infra + apm + profiled + network + fargate")}
      {_drow("Containers",        f"{cx.container_count:,.0f}")}
      {_drow("Host+Container TS", f"{cx.host_container_ts:,.0f}", f"× {TS_PER_HOST} each")}
      {_drow("Custom Metrics",    f"{snap.custom_metrics:,.0f}")}
      {_drow("Total NumSeries",   f"{cx.total_ts:,.0f}", "all sources")}
    </tbody>
  </table>
</section>
<section class="panel">
  <div class="panel-h"><h2>Sizing detail — tracing</h2></div>
  <table class="detail">
    <thead><tr><th>Metric</th><th>Value</th><th>Note</th></tr></thead>
    <tbody>
      {_drow("Ingested spans", f"{cx.ingested_spans_gb_month:,.1f} GB/mo", "twol_ingested_events_bytes")}
      {_drow("Daily ingested", f"{cx.daily_spans_ingest_gb:,.1f} GB/day", "÷ 30")}
      {_drow("→ Monitoring",   f"{cx.daily_spans_mon_gb:,.1f} GB/day", f"{int(SPAN_TIER_MON*100)}%")}
      {_drow("→ Compliance",   f"{cx.daily_spans_comp_gb:,.1f} GB/day", f"{int(SPAN_TIER_COMP*100)}%")}
    </tbody>
  </table>
</section>
<section class="panel">
  <div class="panel-h"><h2>Sizing detail — RUM</h2></div>
  <table class="detail">
    <thead><tr><th>Metric</th><th>Value</th><th>Note</th></tr></thead>
    <tbody>
      {_drow("Sessions/month", f"{cx.rum_sessions_monthly:,.0f}")}
      {_drow("Sessions/day",   f"{cx.rum_sessions_daily:,.0f}", "÷ 30")}
      {_drow("Recordings/day", f"{cx.rum_session_recording_daily:,.0f}", "replay ÷ 30")}
    </tbody>
  </table>
</section>

<p class="assumptions">
  Assumptions: avg log <b>{AVG_LOG_SIZE_KB} KB</b> · avg span <b>{AVG_SPAN_SIZE_KB} KB</b>
  · <b>{TS_PER_HOST}</b> TS/host · <b>{int(DAYS_PER_MONTH)}</b> days/mo
  · logs <b>{int(LOG_TIER_MON*100)}/{int(LOG_TIER_COMP*100)}</b> mon/comp
  · spans <b>{int(SPAN_TIER_MON*100)}/{int(SPAN_TIER_COMP*100)}</b> mon/comp
</p>

</main>
<footer class="ftr">Datadog → Coralogix · {snap.month} · {snap.pulled_at[:10]} · dd_usage_pull.py</footer>
</body>
</html>"""
    return html.encode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 10 — ZIP packaging
# ═══════════════════════════════════════════════════════════════════════════

def build_zip(
    label: str,
    raw_json: bytes,
    per_month_csvs: list[tuple[str, bytes]],
    xlsx_data: bytes | None,
    html_data: bytes,
    out_dir: Path,
) -> Path:
    zip_name = f"datadog_coralogix_report_{label}.zip"
    zip_path  = out_dir / zip_name

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"datadog_raw_{label}.json",         raw_json)
        for month, csv_bytes in per_month_csvs:
            zf.writestr(f"datadog_usage_{month}.csv",    csv_bytes)
        zf.writestr(f"report_{label}.html",              html_data)
        if xlsx_data:
            zf.writestr(f"coralogix_sizing_{label}.xlsx", xlsx_data)

    return zip_path


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 11 — Main
# ═══════════════════════════════════════════════════════════════════════════

def _previous_month() -> str:
    today = datetime.now(timezone.utc)
    first = today.replace(day=1)
    last_month = first - timedelta(days=1)
    return last_month.strftime("%Y-%m")


def compute_month_range(end_month: str, lookback: int = 2) -> list[str]:
    """Return a list of YYYY-MM strings from (end_month - lookback) through end_month."""
    year, mon = map(int, end_month.split("-"))
    months = []
    for i in range(lookback, -1, -1):
        m = mon - i
        y = year
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y:04d}-{m:02d}")
    return months


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pull Datadog usage and convert to Coralogix sizing.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--month", default=os.getenv("DD_MONTH", _previous_month()),
                        help="Target month YYYY-MM (default: previous month)")
    parser.add_argument("--site",  default=os.getenv("DD_SITE", "datadoghq.com"),
                        help="Datadog site (default: datadoghq.com)")
    parser.add_argument("--out",   default=".", help="Output directory (default: current dir)")
    args = parser.parse_args()

    month   = args.month
    site    = args.site
    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    api_key = os.getenv("DD_API_KEY", "").strip()
    app_key = os.getenv("DD_APP_KEY", os.getenv("DD_APPLICATION_KEY", "")).strip()

    if not api_key:
        sys.exit("\n  DD_API_KEY is not set. Add it to your .env file or environment.\n")
    if not app_key:
        sys.exit("\n  DD_APP_KEY is not set. Add it to your .env file or environment.\n")

    # Compute the 3-month window (target month + 2 prior months)
    months = compute_month_range(month, lookback=2)
    start_month = months[0]
    label = f"{months[0]}_to_{months[-1]}"

    print(f"\n  Datadog → Coralogix Usage Report")
    print(f"  Site    : {site}")
    print(f"  Range   : {months[0]} → {months[-1]}  ({len(months)} months)")
    print(f"  Output  : {out_dir}\n")

    client = DatadogClient(api_key, app_key, site)

    # ── Fetch usage data (single endpoint — volumes only, no cost/billing) ─
    summary_raw = None

    print(f"  Fetching usage summary ({start_month} → {month}) …")
    try:
        summary_raw = client.usage_summary(start_month, month)
        n_months_returned = len(summary_raw.get("usage", []))
        print(f"        OK — {n_months_returned} month(s) returned")
    except Exception as exc:
        print(f"        WARN: {exc}")
        summary_raw = {}

    # ── Process each month ────────────────────────────────────────────────
    print(f"\n  Processing {len(months)} months …")
    all_pairs: list[tuple[UsageSnapshot, CoralogixSizing]] = []
    all_usage_items = summary_raw.get("usage", []) if summary_raw else []

    for m in months:
        # Filter the full usage list to just this month's entry
        m_items = [
            u for u in all_usage_items
            if str(u.get("date", u.get("start_date", ""))).startswith(m)
        ]
        if not m_items:
            print(f"    {m}: no data returned — skipping")
            continue
        m_raw = {"usage": m_items}
        snap = extract_usage_snapshot(m_raw, m, site)
        cx = compute_coralogix_sizing(snap)
        all_pairs.append((snap, cx))
        print(f"    {m}: ✓  logs={cx.daily_logs_gb:.1f} GB/day  "
              f"metrics={cx.total_ts:,.0f} TS  "
              f"tracing={cx.daily_spans_ingest_gb:.1f} GB/day")

    if not all_pairs:
        sys.exit("\n  No data could be processed. Check your API keys and month range.\n")

    # Latest month is the primary sizing reference
    snap, cx = all_pairs[-1]
    trends = compute_trends(all_pairs)

    # ── Generate outputs ──────────────────────────────────────────────────
    print("\n  Generating outputs …")
    combined_raw = {
        "range": {"start": months[0], "end": months[-1]},
        "months": {s.month: s.raw for s, _ in all_pairs},
    }
    raw_json     = json.dumps(combined_raw, indent=2, default=str).encode()
    per_month_csvs = [(s.month, generate_csv(s, c)) for s, c in all_pairs]
    xlsx_data    = generate_xlsx(snap, cx, all_pairs=all_pairs)
    html_data    = generate_html(snap, cx, all_pairs=all_pairs, trends=trends)

    zip_path = build_zip(label, raw_json, per_month_csvs, xlsx_data, html_data, out_dir)

    # ── Print summary ─────────────────────────────────────────────────────
    print(f"\n  ══════════════════════════════════════════════════")
    print(f"  Coralogix TCO Sizing — Latest Month: {snap.month}")
    print(f"  ══════════════════════════════════════════════════")
    print(f"  Logs     : {cx.daily_logs_gb:.2f} GB/day")
    print(f"  Metrics  : {cx.total_ts:,.0f} NumSeries")
    print(f"  Tracing  : {cx.daily_spans_ingest_gb:.2f} GB/day")
    print(f"  RUM      : {cx.rum_sessions_daily:,.0f} sessions/day  |  {cx.rum_session_recording_daily:,.0f} recordings/day")
    print(f"  Checkly  : {cx.synthetics_api_daily:,.0f} API/day  |  {cx.synthetics_browser_daily:,.0f} browser/day")
    if len(all_pairs) > 1:
        print(f"\n  Month-over-Month Trends (last → current):")
        last_t = trends[-1]
        for key, name in [("logs_pct","Logs"),("metrics_pct","Metrics"),
                           ("tracing_pct","Tracing"),("rum_pct","RUM")]:
            pct = last_t.get(key)
            if pct is not None:
                arrow = "▲" if pct > 0 else "▼"
                print(f"    {name:<10}: {arrow} {abs(pct):.1f}%")
    print(f"  ══════════════════════════════════════════════════")
    print(f"\n  Output ZIP : {zip_path}")
    print(f"  Contents   :")
    print(f"    datadog_raw_{label}.json")
    for m, _ in per_month_csvs:
        print(f"    datadog_usage_{m}.csv")
    if xlsx_data:
        print(f"    coralogix_sizing_{label}.xlsx")
    print(f"    report_{label}.html")
    print()


if __name__ == "__main__":
    main()
